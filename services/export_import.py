"""
Export/Import service for Laboratory Management System
Handles exporting data to Excel, PDF, CSV and importing from Excel, CSV
"""
import logging
from pathlib import Path
from typing import Tuple, List, Dict
import pandas as pd
from datetime import datetime
from database import db_manager
from models import Patient, TreatmentStatus, Gender
from config import config, get_status_display
from services.statistics import statistics_service
import tempfile
from PIL import Image as PILImage


logger = logging.getLogger(__name__)


class ExportImportService:
    """Export/Import service class"""
    
    @staticmethod
    def export_to_excel(output_path: str, filters: dict = None) -> Tuple[bool, str]:
        """
        Экспорт пациентов в Excel файл с красивым форматированием
        
        Args:
            output_path: Путь для сохранения Excel файла
            filters: Опциональные фильтры для применения
            
        Returns:
            Tuple of (success, message)
        """
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
            
            with db_manager.session_scope(expire_on_commit=False) as session:
                query = session.query(Patient)
                
                # Apply filters if provided
                if filters:
                    if 'status' in filters and filters['status']:
                        query = query.filter(Patient.treatment_status == filters['status'])
                    if 'disease_type' in filters and filters['disease_type']:
                        query = query.filter(Patient.disease_type == filters['disease_type'])
                
                patients = query.all()
                
                # Detach patients from session
                for patient in patients:
                    session.expunge(patient)
                
                # Convert to DataFrame
                data = []
                for patient in patients:
                    data.append({
                        'ID': patient.id,
                        'ФИО': patient.full_name,
                        'Телефон': patient.phone,
                        'Год рождения': patient.birth_year,
                        'Пол': 'Мужской' if patient.gender == Gender.MALE else 'Женский',
                        'Адрес': patient.address or '',
                        'Дата регистрации': patient.registration_date.strftime('%d.%m.%Y %H:%M'),
                        'Тип заболевания': patient.disease_type,
                        'Заболевание': patient.disease_name,
                        'Лечащий врач': patient.treating_doctor or '',
                        'Статус лечения': get_status_display(patient.treatment_status),
                        'Следующий прием': patient.next_appointment.strftime('%d.%m.%Y') if patient.next_appointment else '',
                        'Примечание': patient.notes or ''
                    })
                
                df = pd.DataFrame(data)
                
                # Save to Excel
                output_file = Path(output_path)
                output_file.parent.mkdir(parents=True, exist_ok=True)
                
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Пациенты')
                    
                    worksheet = writer.sheets['Пациенты']
                    
                    # Define styles
                    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    header_border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    cell_font = Font(name='Calibri', size=10)
                    cell_border = Border(
                        left=Side(style='thin', color='D0D0D0'),
                        right=Side(style='thin', color='D0D0D0'),
                        top=Side(style='thin', color='D0D0D0'),
                        bottom=Side(style='thin', color='D0D0D0')
                    )
                    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # Apply header styling
                    for col_num, column in enumerate(df.columns, 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = header_border
                        cell.alignment = header_alignment
                    
                    # Apply data cell styling
                    for row_num in range(2, len(df) + 2):
                        for col_num in range(1, len(df.columns) + 1):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.font = cell_font
                            cell.border = cell_border
                            cell.alignment = cell_alignment
                            
                            # Alternate row colors
                            if row_num % 2 == 0:
                                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 3, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Freeze header row
                    worksheet.freeze_panes = 'A2'
                    
                    # Add title
                    worksheet.insert_rows(1)
                    title_cell = worksheet.cell(row=1, column=1)
                    title_cell.value = f"Отчет по пациентам ({datetime.now().strftime('%d.%m.%Y %H:%M')})"
                    title_cell.font = Font(name='Calibri', size=14, bold=True, color='4472C4')
                    worksheet.merge_cells('A1:M1')
                
                logger.info(f"Exported {len(patients)} patients to Excel: {output_path}")
                return True, f"Экспортировано {len(patients)} пациентов"
                
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return False, f"Ошибка экспорта в Excel: {str(e)}"
    
    @staticmethod
    def export_to_csv(output_path: str, filters: dict = None) -> Tuple[bool, str]:
        """
        Экспорт пациентов в CSV файл
        
        Args:
            output_path: Путь для сохранения CSV файла
            filters: Опциональные фильтры для применения
            
        Returns:
            Tuple of (success, message)
        """
        try:
            with db_manager.session_scope(expire_on_commit=False) as session:
                query = session.query(Patient)
                
                # Apply filters if provided
                if filters:
                    if 'status' in filters and filters['status']:
                        query = query.filter(Patient.treatment_status == filters['status'])
                    if 'disease_type' in filters and filters['disease_type']:
                        query = query.filter(Patient.disease_type == filters['disease_type'])
                
                patients = query.all()
                
                # Detach patients from session
                for patient in patients:
                    session.expunge(patient)
                
                # Convert to DataFrame
                data = []
                for patient in patients:
                    data.append({
                        'ID': patient.id,
                        'ФИО': patient.full_name,
                        'Телефон': patient.phone,
                        'Год рождения': patient.birth_year,
                        'Пол': 'Мужской' if patient.gender == Gender.MALE else 'Женский',
                        'Адрес': patient.address or '',
                        'Дата регистрации': patient.registration_date.strftime('%d.%m.%Y %H:%M'),
                        'Тип заболевания': patient.disease_type,
                        'Заболевание': patient.disease_name,
                        'Лечащий врач': patient.treating_doctor or '',
                        'Статус лечения': get_status_display(patient.treatment_status),
                        'Следующий прием': patient.next_appointment.strftime('%d.%m.%Y') if patient.next_appointment else '',
                        'Примечание': patient.notes or ''
                    })
                
                df = pd.DataFrame(data)
                
                # Save to CSV
                output_file = Path(output_path)
                output_file.parent.mkdir(parents=True, exist_ok=True)
                
                df.to_csv(output_file, index=False, encoding='utf-8-sig')
                
                logger.info(f"Exported {len(patients)} patients to CSV: {output_path}")
                return True, f"Экспортировано {len(patients)} пациентов"
                
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            return False, f"Ошибка экспорта в CSV: {str(e)}"
    
    @staticmethod
    def export_to_pdf(output_path: str, filters: dict = None) -> Tuple[bool, str]:
        """
        Экспорт пациентов в PDF файл с красивым форматированием и поддержкой кириллицы
        
        Args:
            output_path: Путь для сохранения PDF файла
            filters: Опциональные фильтры для применения
            
        Returns:
            Tuple of (success, message)
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            import os
            import sys
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            # Register font that supports Cyrillic
            if sys.platform.startswith('win'):
                # Register Arial from Windows fonts
                arial_path = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts', 'arial.ttf')
                if os.path.exists(arial_path):
                    pdfmetrics.registerFont(TTFont('ArialCyr', arial_path))
                    used_font = 'ArialCyr'
                    used_font_bold = 'ArialCyr'
                else:
                    # Fallback to Helvetica
                    used_font = 'Helvetica'
                    used_font_bold = 'Helvetica'
                    logger.warning("Arial font not found, using Helvetica (Cyrillic may not display)")
            elif sys.platform.startswith('linux'):
                # DejaVuSans on Linux
                font_path = '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
                    used_font = 'DejaVuSans'
                    used_font_bold = 'DejaVuSans'
                else:
                    used_font = 'Helvetica'
                    used_font_bold = 'Helvetica'
                    logger.warning("DejaVuSans font not found on Linux, Cyrillic may not display")
            else:  # MacOS
                # Helvetica on MacOS supports Cyrillic
                used_font = 'Helvetica'
                used_font_bold = 'Helvetica'
            
            with db_manager.session_scope(expire_on_commit=False) as session:
                query = session.query(Patient)
                
                # Apply filters if provided
                if filters:
                    if 'status' in filters and filters['status']:
                        query = query.filter(Patient.treatment_status == filters['status'])
                    if 'disease_type' in filters and filters['disease_type']:
                        query = query.filter(Patient.disease_type == filters['disease_type'])
                
                patients = query.all()
                
                # Detach patients from session
                for patient in patients:
                    session.expunge(patient)
                
                # Create PDF
                output_file = Path(output_path)
                output_file.parent.mkdir(parents=True, exist_ok=True)
                
                doc = SimpleDocTemplate(
                    str(output_file), 
                    pagesize=landscape(A4),
                    rightMargin=0.5*inch,
                    leftMargin=0.5*inch,
                    topMargin=0.5*inch,
                    bottomMargin=0.5*inch
                )
                elements = []
                styles = getSampleStyleSheet()
                
                # Custom styles with Cyrillic-safe font
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    textColor=colors.HexColor('#4472C4'),
                    spaceAfter=20,
                    alignment=TA_CENTER,
                    fontName=used_font_bold
                )
                
                subtitle_style = ParagraphStyle(
                    'CustomSubtitle',
                    parent=styles['Normal'],
                    fontSize=10,
                    textColor=colors.grey,
                    spaceAfter=30,
                    alignment=TA_CENTER,
                    fontName=used_font
                )
                
                # Title
                title = Paragraph("Отчет по пациентам", title_style)
                elements.append(title)
                
                # Date
                date_str = datetime.now().strftime('%d.%m.%Y %H:%M')
                date_para = Paragraph(f"Дата формирования: {date_str}<br/>Количество пациентов: {len(patients)}", subtitle_style)
                elements.append(date_para)
                elements.append(Spacer(1, 0.1 * inch))
                
                # Table data with all columns
                table_data = [['ID', 'ФИО', 'Телефон', 'Год рождения', 'Пол', 'Заболевание', 'Статус', 'Врач']]
                
                for patient in patients:
                    table_data.append([
                        str(patient.id),
                        patient.full_name,
                        patient.phone,
                        str(patient.birth_year),
                        'Мужской' if patient.gender == Gender.MALE else 'Женский',
                        patient.disease_name or '-',
                        get_status_display(patient.treatment_status),
                        patient.treating_doctor or '-'
                    ])
                
                # Create table with better column widths
                table = Table(table_data, colWidths=[0.4*inch, 2.2*inch, 1.2*inch, 0.7*inch, 0.7*inch, 2*inch, 1.2*inch, 1.5*inch], repeatRows=1)
                
                # Beautiful table style
                table.setStyle(TableStyle([
                    # Header styling
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), used_font_bold),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('TOPPADDING', (0, 0), (-1, 0), 12),
                    
                    # Data row styling
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 1), (-1, -1), used_font),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                    
                    # Alternate row colors
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
                    
                    # Grid
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D0D0')),
                    ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#4472C4')),
                ]))
                
                elements.append(table)
                
                # Add footer
                elements.append(Spacer(1, 0.3 * inch))
                footer_style = ParagraphStyle(
                    'Footer',
                    parent=styles['Normal'],
                    fontSize=8,
                    textColor=colors.grey,
                    alignment=TA_CENTER,
                    fontName=used_font
                )
                footer = Paragraph(f"Система управления лабораторией • {config.APP_NAME} v{config.APP_VERSION}", footer_style)
                elements.append(footer)
                
                # Build PDF
                doc.build(elements)
                
                logger.info(f"Exported {len(patients)} patients to PDF: {output_path}")
                return True, f"Экспортировано {len(patients)} пациентов"
                
        except Exception as e:
            logger.error(f"Error exporting to PDF: {e}")
            return False, f"Ошибка экспорта в PDF: {str(e)}"
    
    @staticmethod
    def import_from_excel(input_path: str, user_id: int, update_existing: bool = False) -> Tuple[bool, str, int]:
        """
        Импорт пациентов из Excel файла
        
        Args:
            input_path: Путь к Excel файлу
            user_id: ID пользователя выполняющего импорт
            update_existing: Обновлять существующих пациентов или пропускать их
            
        Returns:
            Tuple of (success, message, imported_count)
        """
        try:
            input_file = Path(input_path)
            
            if not input_file.exists():
                return False, "Файл не найден", 0
            
            # Read Excel file
            df = pd.read_excel(input_file)
            
            # Required columns
            required_columns = ['ФИО', 'Телефон', 'Год рождения', 'Тип заболевания', 'Заболевание']
            
            # Check for required columns
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return False, f"Отсутствуют обязательные колонки: {', '.join(missing_columns)}", 0
            
            imported_count = 0
            errors = []
            
            with db_manager.session_scope() as session:
                for index, row in df.iterrows():
                    try:
                        # Map Excel columns to model fields
                        patient_data = {
                            'full_name': str(row['ФИО']).strip(),
                            'phone': str(row['Телефон']).strip(),
                            'birth_year': int(row['Год рождения']),
                            'gender': Gender.MALE if str(row.get('Пол', '')).strip() == 'Мужской' else Gender.FEMALE,
                            'address': str(row.get('Адрес', '')).strip() if pd.notna(row.get('Адрес')) else None,
                            'disease_type': str(row['Тип заболевания']).strip(),
                            'disease_name': str(row['Заболевание']).strip(),
                            'treating_doctor': str(row.get('Лечащий врач', '')).strip() if pd.notna(row.get('Лечащий врач')) else None,
                            'notes': str(row.get('Примечание', '')).strip() if pd.notna(row.get('Примечание')) else None,
                        }
                        
                        # Parse status if provided
                        status_str = str(row.get('Статус лечения', '')).strip() if pd.notna(row.get('Статус лечения')) else ''
                        if status_str:
                            for status in TreatmentStatus:
                                if status_str in get_status_display(status):
                                    patient_data['treatment_status'] = status
                                    break
                            else:
                                patient_data['treatment_status'] = TreatmentStatus.IN_PROGRESS
                        else:
                            patient_data['treatment_status'] = TreatmentStatus.IN_PROGRESS
                        
                        # Parse next appointment if provided
                        appointment_str = str(row.get('Следующий прием', '')).strip() if pd.notna(row.get('Следующий прием')) else ''
                        if appointment_str:
                            try:
                                patient_data['next_appointment'] = datetime.strptime(appointment_str, '%d.%m.%Y')
                            except:
                                pass
                        
                        # Check if patient already exists
                        existing = session.query(Patient).filter(
                            Patient.phone == patient_data['phone']
                        ).first()
                        
                        if existing:
                            if update_existing:
                                # Update existing patient
                                for field, value in patient_data.items():
                                    setattr(existing, field, value)
                                existing.updated_at = datetime.utcnow()
                                imported_count += 1
                            else:
                                # Skip existing patient
                                continue
                        else:
                            # Create new patient
                            patient_data['created_by'] = user_id
                            patient = Patient(**patient_data)
                            session.add(patient)
                            imported_count += 1
                        
                    except Exception as e:
                        errors.append(f"Строка {index + 1}: {str(e)}")
                        continue
                
                if errors:
                    logger.warning(f"Import completed with errors: {errors}")
                    return True, f"Импортировано {imported_count} пациентов. Ошибки: {len(errors)}", imported_count
                else:
                    logger.info(f"Imported {imported_count} patients from Excel")
                    return True, f"Успешно импортировано {imported_count} пациентов", imported_count
                
        except Exception as e:
            logger.error(f"Error importing from Excel: {e}")
            return False, f"Ошибка импорта из Excel: {str(e)}", 0
    
    @staticmethod
    def import_from_csv(input_path: str, user_id: int, update_existing: bool = False) -> Tuple[bool, str, int]:
        """
        Import patients from CSV file
        
        Args:
            input_path: Path to the CSV file
            user_id: ID of user performing the import
            update_existing: Whether to update existing patients or skip them
            
        Returns:
            Tuple of (success, message, imported_count)
        """
        try:
            input_file = Path(input_path)
            
            if not input_file.exists():
                return False, "Файл не найден", 0
            
            # Read CSV file
            df = pd.read_csv(input_file, encoding='utf-8-sig')
            
            # Required columns
            required_columns = ['ФИО', 'Телефон', 'Год рождения', 'Тип заболевания', 'Заболевание']
            
            # Check for required columns
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return False, f"Отсутствуют обязательные колонки: {', '.join(missing_columns)}", 0
            
            imported_count = 0
            errors = []
            
            with db_manager.session_scope() as session:
                for index, row in df.iterrows():
                    try:
                        # Map CSV columns to model fields
                        patient_data = {
                            'full_name': str(row['ФИО']).strip(),
                            'phone': str(row['Телефон']).strip(),
                            'birth_year': int(row['Год рождения']),
                            'gender': Gender.MALE if str(row.get('Пол', '')).strip() == 'Мужской' else Gender.FEMALE,
                            'address': str(row.get('Адрес', '')).strip() if pd.notna(row.get('Адрес')) else None,
                            'disease_type': str(row['Тип заболевания']).strip(),
                            'disease_name': str(row['Заболевание']).strip(),
                            'treating_doctor': str(row.get('Лечащий врач', '')).strip() if pd.notna(row.get('Лечащий врач')) else None,
                            'notes': str(row.get('Примечание', '')).strip() if pd.notna(row.get('Примечание')) else None,
                        }
                        
                        # Parse status if provided
                        status_str = str(row.get('Статус лечения', '')).strip() if pd.notna(row.get('Статус лечения')) else ''
                        if status_str:
                            for status in TreatmentStatus:
                                if status_str in get_status_display(status):
                                    patient_data['treatment_status'] = status
                                    break
                            else:
                                patient_data['treatment_status'] = TreatmentStatus.IN_PROGRESS
                        else:
                            patient_data['treatment_status'] = TreatmentStatus.IN_PROGRESS
                        
                        # Parse next appointment if provided
                        appointment_str = str(row.get('Следующий прием', '')).strip() if pd.notna(row.get('Следующий прием')) else ''
                        if appointment_str:
                            try:
                                patient_data['next_appointment'] = datetime.strptime(appointment_str, '%d.%m.%Y')
                            except:
                                pass
                        
                        # Check if patient already exists
                        existing = session.query(Patient).filter(
                            Patient.phone == patient_data['phone']
                        ).first()
                        
                        if existing:
                            if update_existing:
                                # Update existing patient
                                for field, value in patient_data.items():
                                    setattr(existing, field, value)
                                existing.updated_at = datetime.utcnow()
                                imported_count += 1
                            else:
                                # Skip existing patient
                                continue
                        else:
                            # Create new patient
                            patient_data['created_by'] = user_id
                            patient = Patient(**patient_data)
                            session.add(patient)
                            imported_count += 1
                        
                    except Exception as e:
                        errors.append(f"Строка {index + 1}: {str(e)}")
                        continue
                
                if errors:
                    logger.warning(f"Import completed with errors: {errors}")
                    return True, f"Импортировано {imported_count} пациентов. Ошибки: {len(errors)}", imported_count
                else:
                    logger.info(f"Imported {imported_count} patients from CSV")
                    return True, f"Успешно импортировано {imported_count} пациентов", imported_count
                
        except Exception as e:
            logger.error(f"Error importing from CSV: {e}")
            return False, f"Ошибка импорта из CSV: {str(e)}", 0


# Global export/import service instance
export_import_service = ExportImportService()